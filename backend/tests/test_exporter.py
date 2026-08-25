import io
import pandas as pd
import openpyxl
from app.services.exporter import export_to_csv, export_to_xlsx

# 1. Test CSV Raw Bytes do NOT contain UTF-8 BOM (EF BB BF)
def test_csv_export_no_bom():
    questions = [
        {
            "Question": "What is the wavelength λ of light?",
            "Option 1": "500 nm",
            "Option 2": "600 nm",
            "Option 3": "700 nm",
            "Option 4": "800 nm",
            "Correct Answer": "500 nm",
            "Topic": "Optics"
        }
    ]
    cols = ["Question", "Option 1", "Option 2", "Option 3", "Option 4", "Correct Answer", "Topic"]
    buf = export_to_csv(questions, cols)
    raw_bytes = buf.getvalue()

    # Verify that raw bytes do NOT start with EF BB BF
    assert not raw_bytes.startswith(b"\xef\xbb\xbf"), "CSV output must not start with UTF-8 BOM"
    
    # First byte must be the first character of the first header ('Q' = 0x51)
    assert raw_bytes[0:1] == b"Q"
    assert raw_bytes.startswith(b"Question,Option 1,Option 2")

# 2. Test Parsed Header is Exactly "Question" (not ï»¿Question)
def test_csv_first_header_exact_string():
    questions = [
        {"Question": "Calculate ΔH for the reaction at 25°C ± 0.5°C.", "Topic": "Thermodynamics"}
    ]
    cols = ["Question", "Topic"]
    buf = export_to_csv(questions, cols)
    
    # Read back as text
    text_content = buf.getvalue().decode("utf-8")
    first_line = text_content.splitlines()[0]
    headers = [h.strip() for h in first_line.split(",")]
    
    assert headers[0] == "Question", f"First header must be 'Question', got {headers[0]!r}"
    assert headers == cols

    # Read via pandas
    df = pd.read_csv(io.StringIO(text_content))
    assert list(df.columns) == cols
    assert df.columns[0] == "Question"

# 3. Test Full Unicode Character Preservation (No corruption / stripping)
def test_unicode_character_preservation():
    unicode_question = {
        "Question": "If λ = 5000 Å, calculate frequency ν where c = 3 × 10⁸ m/s, θ = 45°, π ≈ 3.14159, μ ≥ 0.5, γ ≤ 1.2, T = 100°C ± 2°C.",
        "Option 1": "H₂O + CO₂ → H₂CO₃",
        "Option 2": "CH₃COOH ⇌ CH₃COO⁻ + H⁺",
        "Option 3": "E = mc² ÷ 2",
        "Option 4": "α + β + γ = 180°",
        "Correct Answer": "H₂O + CO₂ → H₂CO₃",
        "Topic": "Physical Chemistry (λ & μ Studies)"
    }
    cols = ["Question", "Option 1", "Option 2", "Option 3", "Option 4", "Correct Answer", "Topic"]
    buf = export_to_csv([unicode_question], cols)
    
    raw_bytes = buf.getvalue()
    assert not raw_bytes.startswith(b"\xef\xbb\xbf")
    
    decoded_text = raw_bytes.decode("utf-8")
    assert "λ" in decoded_text
    assert "μ" in decoded_text
    assert "γ" in decoded_text
    assert "θ" in decoded_text
    assert "π" in decoded_text
    assert "°" in decoded_text
    assert "±" in decoded_text
    assert "≤" in decoded_text
    assert "≥" in decoded_text
    assert "×" in decoded_text
    assert "÷" in decoded_text
    assert "H₂O" in decoded_text
    assert "CH₃COOH" in decoded_text
    assert "10⁸" in decoded_text
    assert "mc²" in decoded_text

    # Read via pandas to ensure exact match
    df = pd.read_csv(io.StringIO(decoded_text))
    assert df.iloc[0]["Question"] == unicode_question["Question"]
    assert df.iloc[0]["Option 1"] == unicode_question["Option 1"]
    assert df.iloc[0]["Option 2"] == unicode_question["Option 2"]
    assert df.iloc[0]["Option 3"] == unicode_question["Option 3"]
    assert df.iloc[0]["Option 4"] == unicode_question["Option 4"]

# 4. Test Excel XLSX Export Remains Unaffected
def test_xlsx_export_unaffected():
    questions = [
        {"Question": "What is the speed of light?", "Correct Answer": "3 × 10⁸ m/s", "Topic": "Physics"}
    ]
    cols = ["Question", "Correct Answer", "Topic"]
    buf = export_to_xlsx(questions, cols)
    
    # Read back with openpyxl
    wb = openpyxl.load_workbook(buf)
    ws = wb.active
    header_vals = [cell.value for cell in ws[1]]
    assert header_vals == cols
    row2_vals = [cell.value for cell in ws[2]]
    assert row2_vals == ["What is the speed of light?", "3 × 10⁸ m/s", "Physics"]
