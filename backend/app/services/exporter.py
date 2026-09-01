import os
from io import BytesIO
from typing import Any
import pandas as pd
from pathlib import Path
import openpyxl

def get_field_value(q: dict, col: str) -> Any:
    """
    Extracts the correct value for a given export column header from a question object/dict.
    Ensures:
      - Option columns (Answer 1-4, Option 1-4, Option A-D) map strictly to the 4 options.
      - Correct Answer columns map strictly to the authoritative answer key.
      - Other columns (Question, Topic, Subtopic, Difficulty, Score) map to their respective fields.
    """
    if not isinstance(q, dict):
        return ""

    col_clean = "".join(c for c in col.lower() if c.isalnum())

    # Check direct match in q first if present and non-empty
    if col in q and q[col] is not None and str(q[col]).strip() != "":
        options = q.get("options")
        if isinstance(options, list) and len(options) >= 4:
            if col_clean in ("answer1", "option1", "optiona", "choice1", "choicea", "opt1", "opta"):
                return options[0]
            elif col_clean in ("answer2", "option2", "optionb", "choice2", "choiceb", "opt2", "optb"):
                return options[1]
            elif col_clean in ("answer3", "option3", "optionc", "choice3", "choicec", "opt3", "optc"):
                return options[2]
            elif col_clean in ("answer4", "option4", "optiond", "choice4", "choiced", "opt4", "optd"):
                return options[3]
        return q[col]

    data_json = q.get("data_json") if isinstance(q.get("data_json"), dict) else {}
    if col in data_json and data_json[col] is not None and str(data_json[col]).strip() != "":
        return data_json[col]

    options = q.get("options") or data_json.get("options") or []
    if not isinstance(options, list):
        options = []

    # 1. Option 1 / Answer 1
    if col_clean in ("answer1", "option1", "optiona", "choice1", "choicea", "opt1", "opta"):
        if len(options) > 0:
            return options[0]
        for key in ("answer_1", "option_1", "option_a", "Answer 1", "Option 1", "Option A", "choice_1", "choice_a"):
            if key in q and q[key] is not None:
                return q[key]
            if key in data_json and data_json[key] is not None:
                return data_json[key]
        return ""

    # 2. Option 2 / Answer 2
    if col_clean in ("answer2", "option2", "optionb", "choice2", "choiceb", "opt2", "optb"):
        if len(options) > 1:
            return options[1]
        for key in ("answer_2", "option_2", "option_b", "Answer 2", "Option 2", "Option B", "choice_2", "choice_b"):
            if key in q and q[key] is not None:
                return q[key]
            if key in data_json and data_json[key] is not None:
                return data_json[key]
        return ""

    # 3. Option 3 / Answer 3
    if col_clean in ("answer3", "option3", "optionc", "choice3", "choicec", "opt3", "optc"):
        if len(options) > 2:
            return options[2]
        for key in ("answer_3", "option_3", "option_c", "Answer 3", "Option 3", "Option C", "choice_3", "choice_c"):
            if key in q and q[key] is not None:
                return q[key]
            if key in data_json and data_json[key] is not None:
                return data_json[key]
        return ""

    # 4. Option 4 / Answer 4
    if col_clean in ("answer4", "option4", "optiond", "choice4", "choiced", "opt4", "optd"):
        if len(options) > 3:
            return options[3]
        for key in ("answer_4", "option_4", "option_d", "Answer 4", "Option 4", "Option D", "choice_4", "choice_d"):
            if key in q and q[key] is not None:
                return q[key]
            if key in data_json and data_json[key] is not None:
                return data_json[key]
        return ""

    # 5. Correct Answer (Authoritative Answer Key)
    if col_clean in ("correctanswer", "correctoption", "authoritativeanswer", "answerkey", "key", "answer"):
        for key in (
            "correct_answer", "Correct Answer", "final_answer", "final_answer_text",
            "final_answer_key", "source_answer_key", "source_answer_text", "correct_option", "answer"
        ):
            if key in q and q[key] is not None and str(q[key]).strip() != "":
                return q[key]
            if key in data_json and data_json[key] is not None and str(data_json[key]).strip() != "":
                return data_json[key]
        return ""

    # 6. Question Stem
    if col_clean in ("question", "questiontext", "problemstatement", "stem", "prompt"):
        for key in ("question", "Question", "question_text", "Question Text", "problem_statement", "stem", "prompt"):
            if key in q and q[key] is not None:
                return q[key]
            if key in data_json and data_json[key] is not None:
                return data_json[key]
        return ""

    # 7. Topic
    if col_clean in ("topic", "questiontopic", "subject"):
        for key in ("topic", "Topic", "question_topic", "Question Topic", "subject", "Subject"):
            if key in q and q[key] is not None:
                return q[key]
            if key in data_json and data_json[key] is not None:
                return data_json[key]
        return ""

    # 8. Subtopic
    if col_clean in ("subtopic", "sub_topic", "subtopictitle"):
        for key in ("subtopic", "Subtopic", "sub_topic", "Sub Topic"):
            if key in q and q[key] is not None:
                return q[key]
            if key in data_json and data_json[key] is not None:
                return data_json[key]
        return ""

    # 9. Difficulty
    if col_clean in ("difficulty", "difficultylevel", "level"):
        for key in ("difficulty", "Difficulty", "difficulty_level", "Difficulty Level", "level"):
            if key in q and q[key] is not None:
                return q[key]
            if key in data_json and data_json[key] is not None:
                return data_json[key]
        return ""

    # 10. Score
    if col_clean in ("score", "marks", "mark", "points"):
        for key in ("score", "Score", "marks", "Marks", "mark", "points"):
            if key in q and q[key] is not None:
                return q[key]
            if key in data_json and data_json[key] is not None:
                return data_json[key]
        return ""

    # Fallback to case-insensitive key search
    for k, v in q.items():
        if "".join(c for c in str(k).lower() if c.isalnum()) == col_clean:
            return v
    for k, v in data_json.items():
        if "".join(c for c in str(k).lower() if c.isalnum()) == col_clean:
            return v

    return q.get(col, "")


def format_draft_cell_value(val: Any, col: str, metadata: dict | None) -> str:
    """Formats cell values in draft export to clearly indicate MISSING, AI_INFERRED, and REVIEW_REQUIRED states."""
    str_val = str(val or "").strip()
    if not metadata:
        return str_val or "[MISSING]"

    fields_meta = metadata.get("fields", {})
    field_info = fields_meta.get(col, {})
    status = field_info.get("status") or field_info.get("origin")
    review_req = field_info.get("review_required", False)
    ai_sugg = field_info.get("ai_suggestion")

    if not str_val:
        if review_req and ai_sugg:
            return f"[REVIEW_REQUIRED: {ai_sugg}]"
        elif status == "MISSING":
            reason = field_info.get("reason", "")
            return f"[MISSING: {reason}]" if reason else "[MISSING]"
        return "[MISSING]"

    if status == "AI_INFERRED" or field_info.get("origin") == "AI_INFERRED":
        conf = field_info.get("confidence")
        conf_str = f" ({int(conf * 100)}%)" if conf is not None else ""
        return f"{str_val} [AI_INFERRED{conf_str}]"

    return str_val

def export_to_csv(
    questions: list[dict],
    columns: list[str],
    is_draft: bool = False,
    metadata_list: list[dict] | None = None
) -> BytesIO:
    rows = []
    for idx, q in enumerate(questions):
        meta = metadata_list[idx] if metadata_list and idx < len(metadata_list) else None
        if is_draft:
            rows.append({col: format_draft_cell_value(get_field_value(q, col), col, meta) for col in columns})
        else:
            rows.append({col: get_field_value(q, col) for col in columns})
    df = pd.DataFrame(rows, columns=columns)
    buf = BytesIO()
    df.to_csv(buf, index=False, encoding="utf-8")
    buf.seek(0)
    return buf

def export_to_xlsx(
    questions: list[dict],
    columns: list[str],
    template_path: str | None = None,
    sheet_name: str | None = None,
    is_draft: bool = False,
    metadata_list: list[dict] | None = None
) -> BytesIO:
    buf = BytesIO()
    
    # If we have the original template XLSX file and not a draft with custom formatted tags, load to preserve styling
    if not is_draft and template_path and os.path.exists(template_path) and template_path.lower().endswith(('.xlsx', '.xls')):
        try:
            wb = openpyxl.load_workbook(template_path)
            if sheet_name and sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.active
            
            # Clear all rows below header (row 1)
            max_row = ws.max_row
            if max_row > 1:
                ws.delete_rows(2, max_row)
                
            # Write new questions
            for r_idx, q in enumerate(questions, start=2):
                for c_idx, col in enumerate(columns, start=1):
                    val = get_field_value(q, col)
                    ws.cell(row=r_idx, column=c_idx, value=val)
                    
            wb.save(buf)
            buf.seek(0)
            return buf
        except Exception:
            # Fallback to pandas if template loading fails
            pass
            
    # Fallback/standard pandas generation
    rows = []
    for idx, q in enumerate(questions):
        meta = metadata_list[idx] if metadata_list and idx < len(metadata_list) else None
        if is_draft:
            rows.append({col: format_draft_cell_value(get_field_value(q, col), col, meta) for col in columns})
        else:
            rows.append({col: get_field_value(q, col) for col in columns})
            
    df = pd.DataFrame(rows, columns=columns)
    s_name = sheet_name or ("Draft_Review" if is_draft else "Questions")
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=s_name)
        ws = writer.book[s_name]
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = cell.font.copy(bold=True)
    buf.seek(0)
    return buf
